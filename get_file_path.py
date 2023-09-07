import os

from openpyxl.workbook import Workbook


def traverse_files(folder_path):
    file_list = []
    dir_list = []
    file_name_list = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            file_name_list.append(file)
            # 在这里对文件进行操作，例如打印文件路径
            file_list.append(file_path)
        for x in dirs:
            dir_path = os.path.join(root, x)
            dir_file_list = os.listdir(dir_path)
            if len(dir_file_list) == 0:
                dir_list.append(dir_path)
    print(file_list)
    print(dir_list)
    # print(path_list)
    max_len = 0
    for x in file_list:
        path_len = len(x.split("\\"))
        if path_len > max_len:
            max_len = path_len
    all_list = file_list + dir_list
    all_list.sort()
    workbook = Workbook()
    # 获取默认的活动工作表
    sheet = workbook.active
    for index, file_path in enumerate(all_list):
        last_ele = file_path.split("\\")[-1]
        if last_ele in file_name_list:
            sheet.cell(row=index + 1, column=max_len).value = last_ele
            for col_index, x in enumerate(file_path.split("\\")[:-1]):
                sheet.cell(row=index + 1, column=col_index + 1).value = x
        else:
            for col_index, x in enumerate(file_path.split("\\")):
                sheet.cell(row=index + 1, column=col_index + 1).value = x
    # for index, file_path in enumerate(file_list):
    #     last_ele = file_path.split("\\")[-1]
    #     sheet.cell(row=index+1, column=max_len).value = last_ele
    #     for col_index, x in enumerate(file_path.split("\\")[:-1]):
    #         sheet.cell(row=index + 1, column=col_index+1).value = x
    #
    # for index, dir_path in enumerate(dir_list):
    #     for col_index, x in enumerate(dir_path.split("\\")):
    #         sheet.cell(row=index + len(file_list) + 3, column=col_index+1).value = x

    workbook.save("example.xlsx")


# 调用函数遍历文件夹下的所有文件
folder_path = r'D:\xj\666'  # 替换为实际的文件夹路径
traverse_files(folder_path)